from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime
import io
import os
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from src.models.database import db, Profissional, Cidade, Equipamento, Usuario
from src.utils.auditoria import registrar_auditoria

relatorios_bp = Blueprint('relatorios', __name__)

def verificar_permissao_relatorios():
    current_user_id = get_jwt_identity()
    usuario = Usuario.query.get(current_user_id)
    return usuario and usuario.nivel_acesso >= 2  # Editor ou superior

@relatorios_bp.route('/profissionais/pdf', methods=['GET'])
@jwt_required()
def gerar_relatorio_pdf():
    try:
        if not verificar_permissao_relatorios():
            return jsonify({'error': 'Permissão negada'}), 403

        # Obter filtros
        status = request.args.get('status', 'ativo')
        cidade_id = request.args.get('cidade_id')
        equipamento_id = request.args.get('equipamento_id')
        
        # Construir query
        query = Profissional.query
        
        # Aplicar filtros de permissão
        current_user_id = get_jwt_identity()
        usuario = Usuario.query.get(current_user_id)
        if usuario.nivel_acesso < 4 and usuario.cidade_id:
            query = query.filter_by(cidade_id=usuario.cidade_id)
        
        # Aplicar filtros da requisição
        if status == 'ativo':
            query = query.filter_by(ativo=True)
        elif status == 'inativo':
            query = query.filter_by(ativo=False)
        
        if cidade_id:
            query = query.filter_by(cidade_id=int(cidade_id))
        
        if equipamento_id:
            query = query.filter_by(equipamento_id=int(equipamento_id))
        
        profissionais = query.all()
        
        # Criar PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Centralizado
        )
        
        # Conteúdo
        story = []
        
        # Título
        title = Paragraph("Relatório de Profissionais", title_style)
        story.append(title)
        
        # Informações do relatório
        info_data = [
            ['Data de Geração:', datetime.now().strftime('%d/%m/%Y %H:%M')],
            ['Status:', status.title()],
            ['Total de Registros:', str(len(profissionais))]
        ]
        
        if cidade_id:
            cidade = Cidade.query.get(cidade_id)
            info_data.append(['Cidade:', cidade.nome if cidade else 'N/A'])
        
        if equipamento_id:
            equipamento = Equipamento.query.get(equipamento_id)
            info_data.append(['Equipamento:', equipamento.nome if equipamento else 'N/A'])
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Tabela de profissionais
        if profissionais:
            # Cabeçalho da tabela
            data = [['Nome', 'CPF', 'Profissão', 'Cargo', 'Equipamento', 'Status']]
            
            # Dados dos profissionais
            for prof in profissionais:
                equipamento = Equipamento.query.get(prof.equipamento_id)
                data.append([
                    prof.nome_completo,
                    prof.cpf,
                    prof.profissao,
                    prof.cargo,
                    equipamento.nome if equipamento else 'N/A',
                    'Ativo' if prof.ativo else 'Inativo'
                ])
            
            # Criar tabela
            table = Table(data, colWidths=[2*inch, 1.2*inch, 1.5*inch, 1.5*inch, 1.5*inch, 0.8*inch])
            table.setStyle(TableStyle([
                # Cabeçalho
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                
                # Dados
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("Nenhum profissional encontrado com os filtros aplicados.", styles['Normal']))
        
        # Gerar PDF
        doc.build(story)
        buffer.seek(0)
        
        # Registrar auditoria
        registrar_auditoria(
            usuario_id=get_jwt_identity(),
            acao='EXPORT',
            tabela='profissionais',
            registro_id=0,
            dados_novos={'tipo': 'PDF', 'filtros': request.args.to_dict()},
            ip_origem=request.remote_addr
        )
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'relatorio_profissionais_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@relatorios_bp.route('/profissionais/excel', methods=['GET'])
@jwt_required()
def gerar_relatorio_excel():
    try:
        if not verificar_permissao_relatorios():
            return jsonify({'error': 'Permissão negada'}), 403

        # Obter filtros (mesmo código do PDF)
        status = request.args.get('status', 'ativo')
        cidade_id = request.args.get('cidade_id')
        equipamento_id = request.args.get('equipamento_id')
        
        # Construir query
        query = Profissional.query
        
        # Aplicar filtros de permissão
        current_user_id = get_jwt_identity()
        usuario = Usuario.query.get(current_user_id)
        if usuario.nivel_acesso < 4 and usuario.cidade_id:
            query = query.filter_by(cidade_id=usuario.cidade_id)
        
        # Aplicar filtros da requisição
        if status == 'ativo':
            query = query.filter_by(ativo=True)
        elif status == 'inativo':
            query = query.filter_by(ativo=False)
        
        if cidade_id:
            query = query.filter_by(cidade_id=int(cidade_id))
        
        if equipamento_id:
            query = query.filter_by(equipamento_id=int(equipamento_id))
        
        profissionais = query.all()
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Profissionais"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Cabeçalhos
        headers = [
            'Nome Completo', 'CPF', 'RG', 'Data Nascimento', 'Escolaridade',
            'Profissão', 'Cargo', 'Vínculo', 'Telefone', 'Email',
            'Data Início Trabalho', 'Endereço', 'Cidade', 'Equipamento', 'Status'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Dados
        for row, prof in enumerate(profissionais, 2):
            cidade = Cidade.query.get(prof.cidade_id)
            equipamento = Equipamento.query.get(prof.equipamento_id)
            
            data = [
                prof.nome_completo,
                prof.cpf,
                prof.rg,
                prof.data_nascimento.strftime('%d/%m/%Y') if prof.data_nascimento else '',
                prof.escolaridade,
                prof.profissao,
                prof.cargo,
                prof.vinculo_institucional,
                prof.telefone,
                prof.email,
                prof.data_inicio_trabalho.strftime('%d/%m/%Y') if prof.data_inicio_trabalho else '',
                prof.endereco_residencial,
                cidade.nome if cidade else 'N/A',
                equipamento.nome if equipamento else 'N/A',
                'Ativo' if prof.ativo else 'Inativo'
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Registrar auditoria
        registrar_auditoria(
            usuario_id=get_jwt_identity(),
            acao='EXPORT',
            tabela='profissionais',
            registro_id=0,
            dados_novos={'tipo': 'Excel', 'filtros': request.args.to_dict()},
            ip_origem=request.remote_addr
        )
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'relatorio_profissionais_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@relatorios_bp.route('/estatisticas', methods=['GET'])
@jwt_required()
def obter_estatisticas():
    try:
        if not verificar_permissao_relatorios():
            return jsonify({'error': 'Permissão negada'}), 403

        current_user_id = get_jwt_identity()
        usuario = Usuario.query.get(current_user_id)
        
        # Base query com filtros de permissão
        base_query = Profissional.query
        if usuario.nivel_acesso < 4 and usuario.cidade_id:
            base_query = base_query.filter_by(cidade_id=usuario.cidade_id)
        
        # Estatísticas gerais
        total_profissionais = base_query.count()
        profissionais_ativos = base_query.filter_by(ativo=True).count()
        profissionais_inativos = base_query.filter_by(ativo=False).count()
        
        # Estatísticas por equipamento
        from sqlalchemy import func
        stats_equipamentos = db.session.query(
            Equipamento.nome,
            func.count(Profissional.id).label('total'),
            func.sum(func.case([(Profissional.ativo == True, 1)], else_=0)).label('ativos'),
            func.sum(func.case([(Profissional.ativo == False, 1)], else_=0)).label('inativos')
        ).join(Profissional, Equipamento.id == Profissional.equipamento_id)
        
        if usuario.nivel_acesso < 4 and usuario.cidade_id:
            stats_equipamentos = stats_equipamentos.filter(Profissional.cidade_id == usuario.cidade_id)
        
        stats_equipamentos = stats_equipamentos.group_by(Equipamento.nome).all()
        
        # Estatísticas por cidade (apenas para Admin Global)
        stats_cidades = []
        if usuario.nivel_acesso == 4:
            stats_cidades = db.session.query(
                Cidade.nome,
                func.count(Profissional.id).label('total'),
                func.sum(func.case([(Profissional.ativo == True, 1)], else_=0)).label('ativos'),
                func.sum(func.case([(Profissional.ativo == False, 1)], else_=0)).label('inativos')
            ).join(Profissional, Cidade.id == Profissional.cidade_id)\
             .group_by(Cidade.nome).all()
        
        # Estatísticas por profissão
        stats_profissoes = db.session.query(
            Profissional.profissao,
            func.count(Profissional.id).label('total')
        )
        
        if usuario.nivel_acesso < 4 and usuario.cidade_id:
            stats_profissoes = stats_profissoes.filter_by(cidade_id=usuario.cidade_id)
        
        stats_profissoes = stats_profissoes.group_by(Profissional.profissao)\
                                         .order_by(func.count(Profissional.id).desc())\
                                         .limit(10).all()
        
        return jsonify({
            'geral': {
                'total_profissionais': total_profissionais,
                'profissionais_ativos': profissionais_ativos,
                'profissionais_inativos': profissionais_inativos,
                'taxa_atividade': round((profissionais_ativos / total_profissionais * 100) if total_profissionais > 0 else 0, 2)
            },
            'por_equipamento': [
                {
                    'equipamento': stat.nome,
                    'total': int(stat.total),
                    'ativos': int(stat.ativos or 0),
                    'inativos': int(stat.inativos or 0)
                }
                for stat in stats_equipamentos
            ],
            'por_cidade': [
                {
                    'cidade': stat.nome,
                    'total': int(stat.total),
                    'ativos': int(stat.ativos or 0),
                    'inativos': int(stat.inativos or 0)
                }
                for stat in stats_cidades
            ],
            'por_profissao': [
                {
                    'profissao': stat.profissao,
                    'total': int(stat.total)
                }
                for stat in stats_profissoes
            ]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

